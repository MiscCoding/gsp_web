#-*- coding: utf-8 -*-
import csv
import datetime
import openpyxl
from collections import OrderedDict

from flask import request, Response, render_template, Blueprint, json, make_response, g, session
import flask_excel as excel
from GSP_WEB import db_session, login_required
from GSP_WEB.common.util.invalidUsage import InvalidUsage
from GSP_WEB.common.util.logUtil import logUtil
from GSP_WEB.models.Regular_Expression_Management_Model import Regular_Expression_Management
#from GSP_WEB.models.wl_maintenance_period import wl_maintenance_period
from GSP_WEB.views.integratedELK import blueprint_page

@blueprint_page.route('/Regular_Expression_Management', methods=['GET'])
@login_required
def Regular_Expression_Management_List():
    logUtil.addLog(request.remote_addr,1,'ELK > Regular_Expression_Management')
    #whitelist = wl_maintenance_period.query.filter_by(datatype='days').first()
    return render_template('integratedELK/Regular_Expression_Management.html'
                           # ,
                           # whiteListPeriod = whitelist.wl_maintenance_period
                           )



@blueprint_page.route('/Regular_Expression_Management/whitelistPeriodSet', methods=['PUT'])
def setRegular_Expression_Management_DashboardLinkNP():
    # whitelist = wl_maintenance_period.query.filter_by(datatype='days').first()
    # whitelist.wl_maintenance_period = request.form.get('whitelistvalue')
    try:
        db_session.commit()
    except Exception as e:
        db_session.rollback()
        raise InvalidUsage('DB 저장 오류', status_code=501)

    return json.dumps({'success':True}), 200, {'ContentType':'application/json'}


@blueprint_page.route('/Regular_Expression_Management/list',methods=['POST'] )
def Regular_Expression_Management_getlist():
    per_page = int(request.form.get('perpage'))
    draw = int(request.form.get('draw'))
    start_idx = int(request.form.get('start'))
    keyword = request.form.get('search_keyword').strip()
    search_keyword_type = request.form.get('search_keyword_type')

    query = Regular_Expression_Management.query

    if keyword != "" and search_keyword_type == "Regular_Exp_Name":
        query = query.filter(Regular_Expression_Management.Regular_Exp_Name.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "Regular_Exp":
        query = query.filter(Regular_Expression_Management.Regular_Exp.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "Description":
        query = query.filter(Regular_Expression_Management.Description.like('%' + keyword + '%'))

    # if keyword != "" and search_keyword_type == "type":
    #     query = query.filter(Regular_Expression_Management.type.like('%' + keyword + '%'))

    # if keyword != "" and search_keyword_type == "Description":
    #     query = query.filter(Regular_Expression_Management.Description.like('%' + keyword + '%'))


    curpage = int(start_idx / per_page) + 1
    cncList = query.order_by(Regular_Expression_Management.seq.desc()).paginate(curpage, per_page, error_out=False)

    result = dict()
    result["draw"] = str(draw)
    result["recordsTotal"] = str(cncList.total)
    result["recordsFiltered"] = str(cncList.total)
    result["data"] = Regular_Expression_Management.serialize_list(cncList.items)
    str_json = json.dumps(result)
    return Response(str_json, mimetype='application/json')

@blueprint_page.route('/regular-expression-list', methods=['POST'])
#@login_required
def add_Regular_Expression():
    #exists = Regular_Expression_Management.query.filter_by(ip=request.form['pattern'].strip()).first()
    # if exists is not None:
    #     raise InvalidUsage('중복 IP가 존재합니다.', status_code=501) # duplication is allowed.

    try:
        _pattern = Regular_Expression_Management()
        _pattern.Regular_Exp_Name = request.form['Regular_Exp_Name']
        # _pattern.IP_Address = request.form['IP_Address'].strip()
        _pattern.Regular_Exp = request.form['Regular_Exp'].strip()
        _pattern.Description = request.form['Description'].strip()
        # _pattern.Password = request.form['Password'].strip()
        # _pattern.description = request.form['desc']
        db_session.add(_pattern)
        db_session.commit()
    except Exception as e:
        db_session.rollback()
        raise InvalidUsage('DB 저장 오류', status_code = 501)

    return ""

@blueprint_page.route('/Regular_Expression_Management/<int:seq>', methods=['PUT'])
#@login_required
def editRegular_Expression_Management_url(seq):
    _pattern = db_session.query(Regular_Expression_Management).filter_by(seq=seq).first()
    try:
        _pattern.Regular_Exp_Name = request.form['Regular_Exp_Name'].strip()
        _pattern.Regular_Exp = request.form['Regular_Exp'].strip()
        _pattern.Description = request.form['Description'].strip()
        # _pattern.Field_Name = request.form['Field_Name'].strip()
        # _pattern.IP_Address = request.form['IP_Address'].strip()
        # _pattern.Description = request.form['Description'].strip()
        # _pattern.Password = request.form['Password'].strip()
        # _pattern.description = request.form['desc']
        # _pattern.mod_dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        db_session.commit()
    except Exception as e:
        db_session.rollback()
        raise InvalidUsage('DB 저장 오류', status_code = 501)
    return ""

@blueprint_page.route('/Regular_Expression_Management/<int:seq>', methods=['DELETE'])
#@login_required
def delete_Regular_Expression_Management_url(seq):
    _pattern = db_session.query(Regular_Expression_Management).filter_by(seq=seq).first()
    if _pattern is not None :
        db_session.delete(_pattern)
        db_session.commit()
    return ""


@blueprint_page.route('/Regular_Expression_Management/uploadlist', methods=['POST'])
def addRegular_Expression_ManagementFileData():
    # jsondata = request.form.get("dna_config");

    # 파일 로드

    file = request.files['file']
    if file.filename.split(".")[1] == 'csv':
        try:
            datalist = []
            spamreader = csv.reader(file)
            next(spamreader)
            for index, row in enumerate(spamreader):

                # if str(row[0]) not in ["Portal", "Video", "AntiVirus", "SNS", "Network", "Server", "Etc"]:
                #     raise InvalidUsage('Type value is not valid', status_code=501)
                #
                # if int(row[2]) not in [8, 16, 24, 32]:
                #     raise InvalidUsage('Mask value is not valid', status_code=501)


                try:
                    _pattern = Regular_Expression_Management()
                    _pattern.Regular_Exp_Name = row[0]
                    _pattern.Regular_Exp = row[1]
                    _pattern.Description = row[2]
                    # _pattern.Description = row[2]
                    # _pattern.Password = row[3]
                    # _pattern.description = row[4]


                    db_session.add(_pattern)
                    db_session.commit()
                except Exception as e:
                    db_session.rollback()
                    raise InvalidUsage('DB 저장 오류' + index + " line ", status_code=501)

        except Exception as e:
            raise InvalidUsage('CSV 로딩 실패, ' + e.message, status_code=501)

    elif file.filename.split(".")[1] == 'xlsx':
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            for index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):

                # if row[0].value not in ["Portal", "Video", "AntiVirus", "SNS", "Network", "Server", "Etc"]:
                #     raise InvalidUsage('Type value is not valid', status_code=501)
                #
                # if int(row[2].value) not in [8, 16, 24, 32]:
                #     raise InvalidUsage('Mask value is not valid', status_code=501)

                try:
                    _pattern = Regular_Expression_Management()
                    _pattern.Regular_Exp_Name = row[0]
                    _pattern.Regular_Exp = row[1]
                    _pattern.Description = row[2]
                    # _pattern.Description = row[2].value
                    # _pattern.Password = row[3].value
                    # _pattern.description = row[4].value

                    db_session.add(_pattern)
                    db_session.commit()
                except Exception as e:
                    db_session.rollback()
                    raise InvalidUsage('DB 저장 오류' + row[0].row + " line ", status_code=501)

        except Exception as e:
            raise InvalidUsage('xlsx 로딩 실패, ' + e.message, status_code=501)


    else:
        raise InvalidUsage('지원하지 않는 파일 포멧 입니다.', status_code=501)

    return json.dumps({'success':True}), 200, {'ContentType':'application/json'}


@blueprint_page.route('/Regular_Expression_Management/excel-list', methods=['GET','POST'])
#@login_required
def getRegular_Expression_Management_ListExcel_url():


    per_page = int(request.form.get('perpage'))
    #draw = int(request.form.get('draw'))
    start_idx = int(request.form.get('start'))
    keyword = request.form.get('search_keyword').strip()
    search_keyword_type = request.form.get('search_keyword_type')

    query = Regular_Expression_Management.query

    # if keyword != "" and search_keyword_type == "IP":
    #     query = query.filter(Regular_Expression_Management.ip.like('%'+keyword+'%'))
    #
    # if keyword != "" and search_keyword_type == "url":
    #     query = query.filter(Regular_Expression_Management.url.like('%' + keyword + '%'))
    if keyword != "" and search_keyword_type == "Regular_Exp_Name":
        query = query.filter(Regular_Expression_Management.Regular_Exp_Name.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "Regular_Exp":
        query = query.filter(Regular_Expression_Management.Regular_Exp.like('%'+keyword+'%'))

    # if keyword != "" and search_keyword_type == "IP_Address":
    #     query = query.filter(Regular_Expression_Management.IP_Address.like('%' + keyword + '%'))

    # if keyword != "" and search_keyword_type == "type":
    #     query = query.filter(Regular_Expression_Management.type.like('%' + keyword + '%'))

    if keyword != "" and search_keyword_type == "Description":
        query = query.filter(Regular_Expression_Management.Description.like('%' + keyword + '%'))

    curpage = int(start_idx / per_page) + 1
    rowCount = query.count()
    # if rowCount > 10000:
    #     rowCount = 10000
    cncList = query.order_by(Regular_Expression_Management.seq.asc()).paginate(curpage, rowCount, error_out=False)
    #inchan = cncList.items[0].ip




    result = OrderedDict()

    result['Regular_Exp_Name'] = list()
    result['Regular_Exp'] = list()
    # result['IP_Address'] = list()
    result['Description'] = list()
    # result['mask'] = list()
    # result['description'] = list()
    # result['url'] = list()
    # result['수정일'] = list()

    for _item in cncList.items:
        result['Regular_Exp_Name'].append(_item.Regular_Exp_Name)
        result['Regular_Exp'].append(_item.Regular_Exp)
        # result['IP_Address'].append(_item.IP_Address)
        result['Description'].append(_item.Description)
        # result['mask'].append(_item.mask)
        # result['url'].append(_item.url)
        # result['description'].append(_item.description)
        # result['수정일'].append(_item.mod_dt)


    return excel.make_response_from_dict(result, "xlsx",
                                          file_name="export_data")

@blueprint_page.route('/Regular_Expression_Management/sample-excel-list', methods=['GET','POST'])
#@login_required
def getReg_Exp_sample():

    sample = request.form.get('sample').strip()

    result = OrderedDict()

    result['Regular_Exp_Name'] = list()
    result['Regular_Exp'] = list()
    # result['IP_Address'] = list()
    result['Description'] = list()
    # result['url'] = list()
    # result['description'] = list()




    result['Regular_Exp_Name'].append("Network")
    # result['IP_Address'].append("111.112.33.54")
    result['Regular_Exp'].append("32")
    result['Description'].append("35")
    # result['url'].append("http://www.daum.net")
    # result['description'].append("자동 등록")


    return excel.make_response_from_dict(result, "xlsx",
                                          file_name="export_data")