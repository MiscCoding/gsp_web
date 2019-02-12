#-*- coding: utf-8 -*-
import csv
import datetime
import openpyxl
from collections import OrderedDict

from flask import request, Response, render_template, Blueprint, json, make_response, g, session
import flask_excel as excel
from GSP_WEB import db_session, login_required
from GSP_WEB.common.util.invalidUsage import InvalidUsage
from GSP_WEB.common.util.logUtil import logUtil
#from GSP_WEB.models.Customer_Management_Model import Integrated_Customer_Management
#from GSP_WEB.models.Customer_Category_Management_Model import Integrated_Customer_Category
#Model for the setting data import for this page.
from GSP_WEB.models.Type_Category_Management_Model import Integrated_Type_Category
from GSP_WEB.models.IP_Category_Management_Model import Integrated_IP_Category
from GSP_WEB.models.Regular_Expression_Management_Model import Regular_Expression_Management

#Model to store and retrieve the major data.
from GSP_WEB.models.Inter_Operation_Policy_Model import Integrated_Inter_Operation_Policy
from GSP_WEB.models.wl_maintenance_period import wl_maintenance_period
from GSP_WEB.views.integratedELK import blueprint_page

@blueprint_page.route('/Inter_Operation_Policy', methods=['GET'])
@login_required
def InterOperation_List():
    logUtil.addLog(request.remote_addr,1,'ELK > Inter_Operation_Policy')
    whitelist = wl_maintenance_period.query.filter_by(datatype='days').first()


    typeCategory = db_session.query(Integrated_Type_Category.Field_Name.distinct().label("Field_Name"))
    typeCategoryList = [row.Field_Name for row in typeCategory.all()]

    IPCategory = db_session.query(Integrated_IP_Category.Field_Name.distinct().label("Field_Name"))
    IPCategoryList = [row.Field_Name for row in IPCategory.all()]

    RegularExpression = db_session.query(Regular_Expression_Management.Regular_Exp_Name.distinct().label("Regular_Exp_Name"))
    RegularExpressionList = [row.Regular_Exp_Name for row in RegularExpression.all()]


    return render_template('integratedELK/Inter_Operation_Policy.html', typeCategoryList = typeCategoryList, IPCategoryList = IPCategoryList, RegularExpressionList = RegularExpressionList)



@blueprint_page.route('/Inter_Operation_Policy/whitelistPeriodSet', methods=['PUT'])
def setInterOperation_DashboardLinkNP():
    whitelist = wl_maintenance_period.query.filter_by(datatype='days').first()
    whitelist.wl_maintenance_period = request.form.get('whitelistvalue')
    try:
        db_session.commit()
    except Exception as e:
        db_session.rollback()
        raise InvalidUsage('DB 저장 오류', status_code=501)

    return json.dumps({'success':True}), 200, {'ContentType':'application/json'}


@blueprint_page.route('/Inter_Operation_Policy/list',methods=['POST'] )
def InterOperation_getlist():
    per_page = int(request.form.get('perpage'))
    draw = int(request.form.get('draw'))
    start_idx = int(request.form.get('start'))
    keyword = request.form.get('search_keyword').strip()
    search_keyword_type = request.form.get('search_keyword_type')

    query = Integrated_Inter_Operation_Policy.query

    if keyword != "" and search_keyword_type == "Type":
        query = query.filter(Integrated_Inter_Operation_Policy.Type.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "IPS_Policy":
        query = query.filter(Integrated_Inter_Operation_Policy.IPS_Policy.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "IPS_Policy_No":
        query = query.filter(Integrated_Inter_Operation_Policy.IPS_Policy_No.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "SRC_IP_Type":
        query = query.filter(Integrated_Inter_Operation_Policy.SRC_IP_Type.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "DST_IP_Type":
        query = query.filter(Integrated_Inter_Operation_Policy.DST_IP_Type.like('%' + keyword + '%'))

    if keyword != "" and search_keyword_type == "Regular_Exp_Name":
        query = query.filter(Integrated_Inter_Operation_Policy.Regular_Exp_Name.like('%' + keyword + '%'))

    # if keyword != "" and search_keyword_type == "type":
    #     query = query.filter(Integrated_Inter_Operation_Policy.type.like('%' + keyword + '%'))

    # if keyword != "" and search_keyword_type == "Description":
    #     query = query.filter(Integrated_Inter_Operation_Policy.Description.like('%' + keyword + '%'))


    curpage = int(start_idx / per_page) + 1
    cncList = query.order_by(Integrated_Inter_Operation_Policy.seq.desc()).paginate(curpage, per_page, error_out=False)

    result = dict()
    result["draw"] = str(draw)
    result["recordsTotal"] = str(cncList.total)
    result["recordsFiltered"] = str(cncList.total)
    result["data"] = Integrated_Inter_Operation_Policy.serialize_list(cncList.items)
    str_json = json.dumps(result)
    return Response(str_json, mimetype='application/json')

@blueprint_page.route('/inter-operation-policy-list', methods=['POST'])
#@login_required
def add_InterOperation_Element():
    #exists = Integrated_Inter_Operation_Policy.query.filter_by(ip=request.form['pattern'].strip()).first()
    # if exists is not None:
    #     raise InvalidUsage('중복 IP가 존재합니다.', status_code=501) # duplication is allowed.

    try:
        _pattern = Integrated_Inter_Operation_Policy()
        _pattern.Type = request.form['Type']
        _pattern.IPS_Policy = request.form['IPS_Policy']
        _pattern.IPS_Policy_No = request.form['IPS_Policy_No']
        _pattern.SRC_IP_Type = request.form['SRC_IP_Type']
        _pattern.DST_IP_Type = request.form['DST_IP_Type']
        _pattern.Regular_Exp_Name = request.form['Regular_Exp_Name']
        # _pattern.IP_Address = request.form['IP_Address'].strip()
        # _pattern.Description = request.form['Description'].strip()
        # _pattern.Password = request.form['Password'].strip()
        # _pattern.description = request.form['desc']
        db_session.add(_pattern)
        db_session.commit()
    except Exception as e:
        db_session.rollback()
        raise InvalidUsage('DB 저장 오류', status_code = 501)

    return ""

@blueprint_page.route('/Inter_Operation_Policy/<int:seq>', methods=['PUT'])
#@login_required
def editInterOperation_Element_url(seq):
    _pattern = db_session.query(Integrated_Inter_Operation_Policy).filter_by(seq=seq).first()
    try:
        _pattern.Type = request.form['Type'].strip()
        _pattern.IPS_Policy = request.form['IPS_Policy'].strip()
        _pattern.IPS_Policy_No = request.form['IPS_Policy_No'].strip()
        _pattern.SRC_IP_Type = request.form['SRC_IP_Type'].strip()
        _pattern.DST_IP_Type = request.form['DST_IP_Type'].strip()
        _pattern.Regular_Exp_Name = request.form['Regular_Exp_Name'].strip()


        # _pattern.Customer_Category = request.form['Customer_Category'].strip()
        # _pattern.Customer_Name = request.form['Customer_Name'].strip()
        # _pattern.IP_Address = request.form['IP_Address'].strip()
        # _pattern.Branch = request.form['Branch'].strip()
        # # _pattern.IP_Address = request.form['IP_Address'].strip()
        # _pattern.Description = request.form['Description'].strip()
        # _pattern.Password = request.form['Password'].strip()
        # _pattern.description = request.form['desc']
        # _pattern.mod_dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        db_session.commit()
    except Exception as e:
        db_session.rollback()
        raise InvalidUsage('DB 저장 오류', status_code = 501)
    return ""

@blueprint_page.route('/Inter_Operation_Policy/<int:seq>', methods=['DELETE'])
#@login_required
def delete_InterOperation_element_url(seq):
    _pattern = db_session.query(Integrated_Inter_Operation_Policy).filter_by(seq=seq).first()
    if _pattern is not None :
        db_session.delete(_pattern)
        db_session.commit()
    return ""


@blueprint_page.route('/Inter_Operation_Policy/uploadlist', methods=['POST'])
def addInterOperation_ElementFileData():
    # jsondata = request.form.get("dna_config");

    # 파일 로드

    file = request.files['file']
    if file.filename.split(".")[1] == 'csv':
        try:
            datalist = []
            spamreader = csv.reader(file)
            next(spamreader)
            for index, row in enumerate(spamreader):

                # if str(row[0]) not in ["Portal", "Video", "AntiVirus", "SNS", "Network", "Server", "Etc"]:
                #     raise InvalidUsage('Type value is not valid', status_code=501)
                #
                # if int(row[2]) not in [8, 16, 24, 32]:
                #     raise InvalidUsage('Mask value is not valid', status_code=501)


                try:
                    _pattern = Integrated_Inter_Operation_Policy()
                    _pattern.Type = row[0]
                    _pattern.IPS_Policy = row[1]
                    _pattern.IP_Address = row[2]
                    _pattern.Branch = row[3]
                    _pattern.Description = row[4]
                    # _pattern.Description = row[2]
                    # _pattern.Password = row[3]
                    # _pattern.description = row[4]


                    db_session.add(_pattern)
                    db_session.commit()
                except Exception as e:
                    db_session.rollback()
                    raise InvalidUsage('DB 저장 오류' + index + " line ", status_code=501)

        except Exception as e:
            raise InvalidUsage('CSV 로딩 실패, ' + e.message, status_code=501)

    elif file.filename.split(".")[1] == 'xlsx':
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            for index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):

                # if row[0].value not in ["Portal", "Video", "AntiVirus", "SNS", "Network", "Server", "Etc"]:
                #     raise InvalidUsage('Type value is not valid', status_code=501)
                #
                # if int(row[2].value) not in [8, 16, 24, 32]:
                #     raise InvalidUsage('Mask value is not valid', status_code=501)

                try:
                    _pattern = Integrated_Inter_Operation_Policy()
                    _pattern.Customer_Category = row[0].value
                    _pattern.Customer_Name = row[1].value
                    _pattern.IP_Address = row[2].value
                    _pattern.Branch = row[3].value
                    _pattern.Description = row[1].value
                    # _pattern.Description = row[2].value
                    # _pattern.Password = row[3].value
                    # _pattern.description = row[4].value

                    db_session.add(_pattern)
                    db_session.commit()
                except Exception as e:
                    db_session.rollback()
                    raise InvalidUsage('DB 저장 오류' + row[0].row + " line ", status_code=501)

        except Exception as e:
            raise InvalidUsage('xlsx 로딩 실패, ' + e.message, status_code=501)


    else:
        raise InvalidUsage('지원하지 않는 파일 포멧 입니다.', status_code=501)

    return json.dumps({'success':True}), 200, {'ContentType':'application/json'}


@blueprint_page.route('/Inter_Operation_Policy/excel-list', methods=['GET','POST'])
#@login_required
def getInterOperation_Element_ListExcel_url():


    per_page = int(request.form.get('perpage'))
    #draw = int(request.form.get('draw'))
    start_idx = int(request.form.get('start'))
    keyword = request.form.get('search_keyword').strip()
    search_keyword_type = request.form.get('search_keyword_type')

    query = Integrated_Inter_Operation_Policy.query

    # if keyword != "" and search_keyword_type == "IP":
    #     query = query.filter(Integrated_Customer_Category.ip.like('%'+keyword+'%'))
    #
    # if keyword != "" and search_keyword_type == "url":
    #     query = query.filter(Integrated_Customer_Category.url.like('%' + keyword + '%'))
    if keyword != "" and search_keyword_type == "Customer_Category":
        query = query.filter(Integrated_Inter_Operation_Policy.Customer_Category.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "Customer_Name":
        query = query.filter(Integrated_Inter_Operation_Policy.Customer_Name.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "IP_Address":
        query = query.filter(Integrated_Inter_Operation_Policy.IP_Address.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "Branch":
        query = query.filter(Integrated_Inter_Operation_Policy.Branch.like('%'+keyword+'%'))

    # if keyword != "" and search_keyword_type == "IP_Address":
    #     query = query.filter(Integrated_Customer_Category.IP_Address.like('%' + keyword + '%'))

    # if keyword != "" and search_keyword_type == "type":
    #     query = query.filter(Integrated_Customer_Category.type.like('%' + keyword + '%'))

    if keyword != "" and search_keyword_type == "Description":
        query = query.filter(Integrated_Inter_Operation_Policy.Description.like('%' + keyword + '%'))

    curpage = int(start_idx / per_page) + 1
    rowCount = query.count()
    # if rowCount > 10000:
    #     rowCount = 10000
    cncList = query.order_by(Integrated_Inter_Operation_Policy.seq.asc()).paginate(curpage, rowCount, error_out=False)
    #inchan = cncList.items[0].ip




    result = OrderedDict()

    result['Customer_Category'] = list()
    result['Customer_Name'] = list()
    result['IP_Address'] = list()
    result['Branch'] = list()
    # result['IP_Address'] = list()
    result['Description'] = list()
    # result['mask'] = list()
    # result['description'] = list()
    # result['url'] = list()
    # result['수정일'] = list()

    for _item in cncList.items:
        result['Customer_Category'].append(_item.Customer_Category)
        result['Customer_Name'].append(_item.Customer_Name)
        result['IP_Address'].append(_item.IP_Address)
        result['Branch'].append(_item.Branch)
        # result['IP_Address'].append(_item.IP_Address)
        result['Description'].append(_item.Description)
        # result['mask'].append(_item.mask)
        # result['url'].append(_item.url)
        # result['description'].append(_item.description)
        # result['수정일'].append(_item.mod_dt)


    return excel.make_response_from_dict(result, "xlsx",
                                          file_name="export_data")

@blueprint_page.route('/Inter_Operation_Policy/sample-excel-list', methods=['GET','POST'])
#@login_required
def getElement_InterOperation_sample():

    sample = request.form.get('sample').strip()

    result = OrderedDict()


    result['Customer_Category'] = list()
    result['Customer_Name'] = list()
    result['IP_Address'] = list()
    result['Branch'] = list()
    # result['IP_Address'] = list()
    result['Description'] = list()
    # result['url'] = list()
    # result['description'] = list()




    result['Customer_Category'].append("Network")
    result['Customer_Name'].append("name")
    result['IP_Address'].append("Address")
    result['Branch'].append("Branch")
    # result['IP_Address'].append("111.112.33.54")
    result['Description'].append("32")
    # result['url'].append("http://www.daum.net")
    # result['description'].append("자동 등록")


    return excel.make_response_from_dict(result, "xlsx",
                                          file_name="export_data")