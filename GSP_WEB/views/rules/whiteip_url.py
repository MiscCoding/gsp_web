#-*- coding: utf-8 -*-
import csv
import datetime
import openpyxl
from collections import OrderedDict

from flask import request, Response, render_template, Blueprint, json, make_response, g, session
import flask_excel as excel
from GSP_WEB import db_session, login_required
from GSP_WEB.common.util.invalidUsage import InvalidUsage
from GSP_WEB.common.util.logUtil import logUtil
from GSP_WEB.models.Rules_White_IP_URL import Rules_White_IP_URL
from GSP_WEB.models.wl_maintenance_period import wl_maintenance_period
from GSP_WEB.views.rules import blueprint_page

@blueprint_page.route('/ip-url-white-list', methods=['GET'])
@login_required
def whiteip_url_List():
    logUtil.addLog(request.remote_addr,1,'rules > ip-url-white-list')
    whitelist = wl_maintenance_period.query.filter_by(datatype='days').first()
    return render_template('rules/whiteip_list_url.html', whiteListPeriod = whitelist.wl_maintenance_period)



@blueprint_page.route('/ip-url-white-list/whitelistPeriodSet', methods=['PUT'])
def setDashboardLinkNP():
    whitelist = wl_maintenance_period.query.filter_by(datatype='days').first()
    whitelist.wl_maintenance_period = request.form.get('whitelistvalue')
    try:
        db_session.commit()
    except Exception as e:
        db_session.rollback()
        raise InvalidUsage('DB 저장 오류', status_code=501)

    return json.dumps({'success':True}), 200, {'ContentType':'application/json'}


@blueprint_page.route('/ip-url-white-list/list',methods=['POST'] )
def whiteip_url_getlist():
    per_page = int(request.form.get('perpage'))
    draw = int(request.form.get('draw'))
    start_idx = int(request.form.get('start'))
    keyword = request.form.get('search_keyword').strip()
    search_keyword_type = request.form.get('search_keyword_type')
    columnIndex = request.form.get('columnIndex')
    sort_style = request.form.get('sort_style')

    query = Rules_White_IP_URL.query

    if keyword != "" and search_keyword_type == "ip":
        query = query.filter(Rules_White_IP_URL.ip.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "mask":
        query = query.filter(Rules_White_IP_URL.ip.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "url":
        query = query.filter(Rules_White_IP_URL.url.like('%' + keyword + '%'))

    if keyword != "" and search_keyword_type == "type":
        query = query.filter(Rules_White_IP_URL.type.like('%'+keyword+'%'))

    if keyword != "" and search_keyword_type == "description":
        query = query.filter(Rules_White_IP_URL.description.like('%' + keyword + '%'))

    if keyword != "" and search_keyword_type == "cre_dt":
        query = query.filter(Rules_White_IP_URL.cre_dt.like('%' + keyword + '%'))

    if keyword != "" and search_keyword_type == "mod_dt":
        query = query.filter(Rules_White_IP_URL.mod_dt.like('%' + keyword + '%'))


    curpage = int(start_idx / per_page) + 1

    # cncList = query.order_by(Rules_White_IP_URL.cre_dt.desc()).paginate(curpage, per_page, error_out=False)

    if columnIndex == 'url':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.url.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.url.asc()).paginate(curpage, per_page, error_out=False)

    elif columnIndex == 'ip':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.ip.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.ip.asc()).paginate(curpage, per_page, error_out=False)

    elif columnIndex == 'mask':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.mask.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.mask.asc()).paginate(curpage, per_page, error_out=False)

    elif columnIndex == 'type':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.type.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.type.asc()).paginate(curpage, per_page, error_out=False)

    elif columnIndex == 'description':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.description.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.description.asc()).paginate(curpage, per_page, error_out=False)

    elif columnIndex == 'cre_dt':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.cre_dt.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.cre_dt.asc()).paginate(curpage, per_page, error_out=False)

    elif columnIndex == 'mod_dt':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.mod_dt.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.mod_dt.asc()).paginate(curpage, per_page, error_out=False)

    else:
        cncList = query.order_by(Rules_White_IP_URL.cre_dt.desc()).paginate(curpage, per_page, error_out=False)


    result = dict()
    result["draw"] = str(draw)
    result["recordsTotal"] = str(cncList.total)
    result["recordsFiltered"] = str(cncList.total)
    result["data"] = Rules_White_IP_URL.serialize_list(cncList.items)
    str_json = json.dumps(result)
    return Response(str_json, mimetype='application/json')

@blueprint_page.route('/ip-url-white-list', methods=['POST'])
#@login_required
def addwhiteip_url():
    exists = Rules_White_IP_URL.query.filter_by(ip=request.form['pattern'].strip()).first()
    # if exists is not None:
    #     raise InvalidUsage('중복 IP가 존재합니다.', status_code=501) # duplication is allowed.

    try:
        _pattern = Rules_White_IP_URL()
        _pattern.type = request.form['type']
        _pattern.ip = request.form['pattern'].strip()
        _pattern.url = request.form['url'].strip()
        _pattern.mask = request.form['mask'].strip()
        _pattern.description = request.form['desc']
        db_session.add(_pattern)
        db_session.commit()
    except Exception as e:
        db_session.rollback()
        raise InvalidUsage('DB 저장 오류', status_code = 501)

    return ""

@blueprint_page.route('/ip-url-white-list/<int:seq>', methods=['PUT'])
#@login_required
def editwhiteip_url(seq):
    _pattern = db_session.query(Rules_White_IP_URL).filter_by(seq=seq).first()
    try:
        _pattern.type = request.form['type'].strip()
        _pattern.ip = request.form['pattern'].strip()
        _pattern.url = request.form['url'].strip()
        _pattern.mask = request.form['mask'].strip()
        _pattern.description = request.form['desc']
        _pattern.mod_dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        db_session.commit()
    except Exception as e:
        db_session.rollback()
        raise InvalidUsage('DB 저장 오류', status_code = 501)
    return ""

@blueprint_page.route('/ip-url-white-list-delete', methods=['POST'])
#@login_required
def deletewhiteip_url():

    seqList = request.get_json()
    for alist in seqList['deleteItemList']:
        _pattern = db_session.query(Rules_White_IP_URL).filter_by(seq=alist).first()
        if _pattern is not None :
            db_session.delete(_pattern)
            db_session.commit()

    return json.dumps({'success':True}), 200, {'ContentType':'application/json'}


@blueprint_page.route('/ip-url-white-list/uploadlist', methods=['POST'])
def addWhiteListFileData():
    # jsondata = request.form.get("dna_config");

    # 파일 로드

    file = request.files['file']
    if file.filename.split(".")[1] == 'csv':
        try:
            datalist = []
            spamreader = csv.reader(file)
            next(spamreader)
            for index, row in enumerate(spamreader):

                # if str(row[0]) not in ["Portal", "Video", "AntiVirus", "SNS", "Network", "Server", "Etc"]:
                #     raise InvalidUsage('Type value is not valid', status_code=501)
                #
                # if int(row[2]) not in [8, 16, 24, 32]:
                #     raise InvalidUsage('Mask value is not valid', status_code=501)


                try:
                    _pattern = Rules_White_IP_URL()
                    _pattern.type = row[0]
                    _pattern.ip = row[1]
                    _pattern.mask = row[2]
                    _pattern.url = row[3]
                    _pattern.description = row[4]


                    db_session.add(_pattern)
                    db_session.commit()
                except Exception as e:
                    db_session.rollback()
                    raise InvalidUsage('DB 저장 오류' + index + " line ", status_code=501)

        except Exception as e:
            raise InvalidUsage('CSV 로딩 실패, ' + e.message, status_code=501)

    elif file.filename.split(".")[1] == 'xlsx':
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            for index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):

                # if row[0].value not in ["Portal", "Video", "AntiVirus", "SNS", "Network", "Server", "Etc"]:
                #     raise InvalidUsage('Type value is not valid', status_code=501)
                #
                # if int(row[2].value) not in [8, 16, 24, 32]:
                #     raise InvalidUsage('Mask value is not valid', status_code=501)

                try:
                    _pattern = Rules_White_IP_URL()
                    _pattern.type = row[0].value
                    _pattern.ip = row[1].value
                    _pattern.mask = row[2].value
                    _pattern.url = row[3].value
                    _pattern.description = row[4].value

                    db_session.add(_pattern)
                    db_session.commit()
                except Exception as e:
                    db_session.rollback()
                    raise InvalidUsage('DB 저장 오류' + row[0].row + " line ", status_code=501)

        except Exception as e:
            raise InvalidUsage('xlsx 로딩 실패, ' + e.message, status_code=501)


    else:
        raise InvalidUsage('지원하지 않는 파일 포멧 입니다.', status_code=501)

    return json.dumps({'success':True}), 200, {'ContentType':'application/json'}


@blueprint_page.route('/ip-url-white-list/excel-list', methods=['GET','POST'])
#@login_required
def getWhiteListExcel_url():
    per_page = int(request.form.get('perpage'))
    # draw = int(request.form.get('draw'))
    start_idx = int(request.form.get('start'))
    keyword = request.form.get('search_keyword').strip()
    search_keyword_type = request.form.get('search_keyword_type')
    columnIndex = request.form.get('columnIndex')
    sort_style = request.form.get('sort_style')


    # per_page = int(request.form.get('perpage'))
    # #draw = int(request.form.get('draw'))
    # start_idx = int(request.form.get('start'))
    # keyword = request.form.get('search_keyword').strip()
    # search_keyword_type = request.form.get('search_keyword_type')

    query = Rules_White_IP_URL.query

    if keyword != "" and search_keyword_type == "ip":
        query = query.filter(Rules_White_IP_URL.ip.like('%' + keyword + '%'))

    if keyword != "" and search_keyword_type == "mask":
        query = query.filter(Rules_White_IP_URL.ip.like('%' + keyword + '%'))

    if keyword != "" and search_keyword_type == "url":
        query = query.filter(Rules_White_IP_URL.url.like('%' + keyword + '%'))

    if keyword != "" and search_keyword_type == "type":
        query = query.filter(Rules_White_IP_URL.type.like('%' + keyword + '%'))

    if keyword != "" and search_keyword_type == "description":
        query = query.filter(Rules_White_IP_URL.description.like('%' + keyword + '%'))

    if keyword != "" and search_keyword_type == "cre_dt":
        query = query.filter(Rules_White_IP_URL.cre_dt.like('%' + keyword + '%'))

    if keyword != "" and search_keyword_type == "mod_dt":
        query = query.filter(Rules_White_IP_URL.mod_dt.like('%' + keyword + '%'))

    curpage = int(start_idx / per_page) + 1
    rowCount = query.count()
    per_page = rowCount
    # if rowCount > 10000:
    #     rowCount = 10000
    if columnIndex == 'url':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.url.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.url.asc()).paginate(curpage, per_page, error_out=False)

    elif columnIndex == 'ip':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.ip.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.ip.asc()).paginate(curpage, per_page, error_out=False)

    elif columnIndex == 'mask':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.mask.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.mask.asc()).paginate(curpage, per_page, error_out=False)

    elif columnIndex == 'type':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.type.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.type.asc()).paginate(curpage, per_page, error_out=False)

    elif columnIndex == 'description':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.description.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.description.asc()).paginate(curpage, per_page, error_out=False)

    elif columnIndex == 'cre_dt':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.cre_dt.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.cre_dt.asc()).paginate(curpage, per_page, error_out=False)

    elif columnIndex == 'mod_dt':
        if sort_style == 'desc':
            cncList = query.order_by(Rules_White_IP_URL.mod_dt.desc()).paginate(curpage, per_page, error_out=False)
        else:
            cncList = query.order_by(Rules_White_IP_URL.mod_dt.asc()).paginate(curpage, per_page, error_out=False)

    else:
        cncList = query.order_by(Rules_White_IP_URL.cre_dt.desc()).paginate(curpage, per_page, error_out=False)




    result = OrderedDict()

    result['registerDate'] = list()
    result['type'] = list()
    result['ip'] = list()
    result['mask'] = list()
    result['description'] = list()
    result['url'] = list()
    result['modifyDate'] = list()

    for _item in cncList.items:
        result['registerDate'].append(_item.cre_dt)
        result['type'].append(_item.type)
        result['ip'].append(_item.ip)
        result['mask'].append(_item.mask)
        result['url'].append(_item.url)
        result['description'].append(_item.description)
        result['modifyDate'].append(_item.mod_dt)


    return excel.make_response_from_dict(result, "csv",
                                          file_name="export_data")

@blueprint_page.route('/ip-url-white-list/sample-excel-list', methods=['GET','POST'])
#@login_required
def getWhiteListExcel_sample():

    sample = request.form.get('sample').strip()

    result = OrderedDict()


    result['type'] = list()
    result['ip'] = list()
    result['mask'] = list()
    result['url'] = list()
    result['description'] = list()




    result['type'].append("Network")
    result['ip'].append("111.112.33.54")
    result['mask'].append("32")
    result['url'].append("http://www.daum.net")
    result['description'].append("자동 등록")


    return excel.make_response_from_dict(result, "xlsx",
                                          file_name="export_data")